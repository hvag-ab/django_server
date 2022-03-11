import csv,tempfile, zipfile,codecs

import openpyxl
from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils import get_column_letter
from openpyxl.styles import Side, Border, colors, PatternFill, Alignment, Font
from django.http import HttpResponse, StreamingHttpResponse
from typing import List, Union, Optional, Callable, Any, IO, Sequence,Generator
from io import StringIO

DATA = Sequence[Union[dict,Sequence]]
Buffer = Union[bytes, IO, Generator]


def file_response(buffer: Buffer,
                  filename: str,
                  ext: str,
                  block_size: int = 4096) -> HttpResponse:
    if ext == 'xlsx':
        content_type = 'application/msexcel'
    elif ext == 'csv':
        content_type = 'text/csv'
    else:
        content_type = 'application/octet-stream'
    if isinstance(buffer, bytes):
        response = HttpResponse(content=buffer, content_type=content_type)
    else:
        try:
            if hasattr(buffer, 'read'):
                chunk = iter(lambda: buffer.read(block_size), b'')
            else:
                chunk = buffer
            response = StreamingHttpResponse(streaming_content=chunk, content_type=content_type)
        except:
            raise TypeError(f'content 参数 的类型{type(buffer)}暂不支持')
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    response['Content-Disposition'] = f"attachment; filename={filename.encode('utf-8').decode('ISO-8859-1')}.{ext}"
    if ext == 'csv':
        response.charset = 'utf-8-sig'
    return response


def nginx_file_response(nginx_file_path:str,
                  filename: str,
                  ext: str) -> HttpResponse:
    """
    使用sendfile的机制："传统的Web服务器在处理文件下载的时候，总是先读入文件内容到应用程序内存，然后再把内存当中的内容发送给客户端浏览器。这种方式在应付当今大负载网站会消耗更多的服务器资源。
    sendfile是现代操作系统支持的一种高性能网络IO方式，操作系统内核的sendfile调用可以将文件内容直接推送到网卡的buffer当中，从而避免了Web服务器读写文件的开销，实现了“零拷贝”模式
    nginx配置文件：
    # file save path: /var/www/files/myfile.tar.gz
    # When passed URI /protected_files/myfile.tar.gz
    # nginx_file_url = "/protected_files/myfile.tar.gz"
    这样当向django view函数发起request时，django负责对用户权限进行判断或者做些其它事情，
    然后向nginx转发url为/protected_files/filename的请求，nginx服务器负责文件/var/www/protected_files/filename的下载：
    location /protected_files {
        internal; # internal指的是Nginx的内部命令，意思是只有内部请求才能使用的，外部请求返回404
        alias /var/www/files;
    }
    """
    if ext == 'xlsx':
        content_type = 'application/msexcel'
    elif ext == 'csv':
        content_type = 'text/csv'
    else:
        content_type = 'application/octet-stream'
    response = HttpResponse()
    response['Content_Type'] = content_type
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    response['Content-Disposition'] = f"attachment; filename={filename.encode('utf-8').decode('ISO-8859-1')}.{ext}"
    response['X-Accel-Redirect'] = nginx_file_path
    return response


class ExcelToData:

    def __init__(self, file: Buffer, header: Optional[dict] = None, first_row: int = 1,
                 sheet_no: int = 0, func: Optional[Callable[[dict], Any]] = None):
        """
        :param file: 上传的excel二进制数据
        :param header: # excel表头映射成模型的字段名
        :param serializer: # 序列化器 验证excel中的数据
        :param first_row: # 从哪一行开始获取数据
        :param sheet_no: # 获取excel中哪一个sheet
        :param func: # 处理每行数据的额外函数 比如添加额外数据 或者 数据验证 或者 保存数据
        """
        self.wb = openpyxl.load_workbook(file)
        sheet_names = self.wb.sheetnames
        self.worksheet = self.wb[sheet_names[sheet_no]]
        # 获取总行数
        # self.nrows = self.worksheet.max_row
        columns = self.worksheet.max_column
        excel_header = [self.worksheet.cell(row=first_row, column=i).value for i in range(1, columns + 1)]
        if header:
            self.header = [header.get(h, h) for h in excel_header]
        else:
            self.header = excel_header
        self.first_row = first_row
        self.func = func

    @property
    def save(self) -> List[Any]:
        datas: List[Any] = []
        for row in self.worksheet.iter_rows(min_row=self.first_row + 1):
            row_data = [cell.value for cell in row]
            data = dict(zip(self.header, row_data))
            if self.func:
                data = self.func(data)
            datas.append(data)
        self.wb.close()
        return datas


class DataToExcel:

    def __init__(self, data: DATA = None, headers: Union[dict, list] = None, name: str = 'Sheet1',
                 is_header: bool = False,ext: str = 'xlsx', title: str = None, merge: list = None,
                 beauty: bool = True,cell_width=15, need_header=True, skip_rows=0,
                 func: Callable[[dict], Any] = None, header_map:Optional[dict] = None):
        """
        :param headers: excel表格头 {'定义的名称'：’字段名',....}
        :param data: 表格数据 [{'字段名':'字段的值’}]
        :param name: 表格的sheet名
        :param is_header: 是否值导出标题头
        :param func: 映射函数 额外处理data数据
        :param ext: 表格格式后缀 xlsx or xls
        :param title: 表格标题名称
        :param merge: 合并单元格cell  例如['A1:A3'] 合并
        :param beauty: 是否美化格式 当数据量大的时候很费时间
        """
        if data:
            first_data = data[0]
            if isinstance(first_data, dict):
                if not headers:
                    self.headers = dict(zip(first_data.keys(), first_data.keys()))
                else:
                    if not isinstance(headers, dict):
                        raise ValueError('data是List[dict], header 必须是 字典')
                    else:
                        self.headers = headers
            else:
                if not headers:
                    raise ValueError('data是Sequence[Sequence], header 不能为空')
                else:
                    if not isinstance(headers, Sequence):
                        raise ValueError('data是Sequence[Sequence], header 必须 为 Sequence')
                    if len(headers) != len(first_data):
                        raise ValueError('data是Sequence[Sequence], header 长度不匹配')
                    self.headers = headers
        else:
            if not headers:
                raise KeyError('header 和 data 不能同时为空')
            else:
                self.headers = headers
        self.data = data
        if isinstance(self.headers, dict):
            self.excel_field_names = list(self.headers.keys())  # 模型所有字段名
        else:
            self.excel_field_names = self.headers
        self.name = name
        self.is_header = is_header
        self.ext = ext
        self.title = title
        self.merge = merge
        self.beauty = beauty
        self.cell_width = cell_width
        self.need_header = need_header
        self.skip_rows = skip_rows
        self.func = func
        self.header_map = header_map

    @property
    def export(self):

        wb = openpyxl.Workbook()
        ws = wb.active

        if self.title:
            leng = len(self.excel_field_names)
            first_cont = [self.title] + [''] * (leng - 1)
            last_col = get_column_letter(leng)

            ws.append(first_cont)
            ws.merge_cells(f'A1:{last_col}1')
            ws.row_dimensions[1].height = 30
            bold_24_font = Font(name='等线', size=20, italic=False, color='0d0992', bold=True)
            ws['A1'].font = bold_24_font

            headline = 2
        else:
            headline = 1

        if self.need_header:
            if self.header_map:
                excel_field_names = list(map(lambda x:self.header_map.get(x,x), self.excel_field_names))
                ws.append(excel_field_names)
            else:
                ws.append(self.excel_field_names)

        orange_fill = PatternFill(fill_type='solid', fgColor="A6D5B0")
        for ci in range(len(self.excel_field_names)):
            if self.need_header:
                ws.cell(row=headline, column=ci + 1).fill = orange_fill
            col_letter = get_column_letter(ci + 1)
            ws.column_dimensions[col_letter].width = self.cell_width

        if not self.is_header:
            for ind, obj in enumerate(self.data):
                if isinstance(obj, Sequence):
                    data = obj
                elif isinstance(obj, dict):
                    if self.func:
                        obj = self.func(obj)
                    data = [obj.get(v, v) for v in self.headers.values()]
                else:
                    raise ValueError('data 必须是 data是Sequence[Sequence]')
                if ind >= self.skip_rows:
                    ws.append(data)

            if self.merge:
                for mer in self.merge:
                    ws.merge_cells(mer)

        if self.beauty:
            # 数据量大了后 每一cell去遍历 很花时间  注意取舍
            row_max = ws.max_row
            con_max = ws.max_column
            for xi in range(1, row_max + 1):
                for ji in range(1, con_max + 1):
                    ws.cell(row=xi, column=ji).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=xi, column=ji).border = Border(top=Side(border_style='thin', color=colors.BLACK),
                                                               bottom=Side(border_style='thin', color=colors.BLACK),
                                                               left=Side(border_style='thin', color=colors.BLACK),
                                                               right=Side(border_style='thin', color=colors.BLACK))

        tmp = tempfile.NamedTemporaryFile()
        archive = zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED, allowZip64=True)
        writer = ExcelWriter(wb, archive)
        writer.save()
        tmp.seek(0)
        wb.close()
        response = file_response(tmp, self.name, self.ext)
        return response


class Echo:
    """An object that implements just the write method of the file-like
    interface.
    """

    def write(self, value):
        """Write the value by returning it, instead of storing in a buffer."""
        return value


class DataToCSV:

    def __init__(self, data: Union[DATA, "DataFrame"], filename: str = 'somefile'):
        self.data = data
        self.filename = filename

    def gen_data(self, writer):
        for index, data in enumerate(self.data):
            if isinstance(data, Sequence):
                yield writer.writerow(data)

            elif isinstance(data, dict):
                if index == 0:
                    yield writer.writerow(list(data.keys()))
                yield writer.writerow(list(data.values()))

    @property
    def export(self):
        if hasattr(self.data,'to_csv'):
            result = StringIO()
            self.data.to_csv(result,index=False,encoding="utf_8_sig")
            result = result.getvalue().encode("utf_8_sig")
        else:
            pseudo_buffer = Echo()
            writer = csv.writer(pseudo_buffer)
            result = self.gen_data(writer)
        return file_response(result,filename=self.filename,ext='csv')


from django.conf import settings
from pathlib import Path


def zip_files(zipname: str, dir: str = None, file_names: List[str] = None):
    """
    :param zipname: zip文件名
    :param dir: 压缩文件所在的文件夹
    :param file_names: 文件名s
    :return:
    """
    if dir is None:
        dir = settings.MEDIA_ROOT
    else:
        dir = Path(dir)

    temp = tempfile.NamedTemporaryFile()
    archive = zipfile.ZipFile(temp, 'w', zipfile.ZIP_DEFLATED)
    for file in file_names:
        file_p = dir / file
        if not file_p.is_file():
            raise ValueError(f'{str(file_p)} 并不是一个文件')
        archive.write(file_p,arcname=file) # arcname解决不需要保存文件所在的原始路径
    archive.close()
    temp.seek(0)
    return file_response(temp, filename=zipname, ext='zip')
