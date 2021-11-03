import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Side, Border, colors, PatternFill, Alignment, Font
import csv
from django.http import HttpResponse, StreamingHttpResponse
from django.core.files.base import File
from typing import List, Union, Optional, Callable, Any
from _io import _IOBase

buffer = Union[bytes, _IOBase, File]
DATA = Union[List[dict], List[Any]]


def file_response(content: buffer, filename: str, ext: str, block_size=4096) -> HttpResponse:
    if ext == 'xlsx':
        content_type = 'application/msexcel'
    elif ext == 'csv':
        content_type = 'text/csv'
    else:
        content_type = 'application/octet-stream'
    if isinstance(content, bytes):
        response = HttpResponse(content=content, content_type=content_type)
    elif isinstance(content, (_IOBase,File)):
        # File类型是django FileField中的类型需要支持 file_response(obj.doc.file,filename=obj.doc.name,ext='doc')
        if hasattr(content, 'read'):
            chunk = iter(lambda: content.read(block_size), b'')
        else:
            chunk = content
        response = StreamingHttpResponse(streaming_content=chunk, content_type=content_type)
    else:
        raise TypeError(f'content 参数 的类型{type(content)}暂不支持')
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    response[
        'Content-Disposition'] = f"attachment; filename={filename.encode('utf-8').decode('ISO-8859-1')}.{ext}"
    return response


class ExcelToData:

    def __init__(self, file: buffer, header: Optional[dict] = None, first_row: int = 1,
                 sheet_no: int = 0, func: Optional[Callable[[dict], Any]] = None):
        """
        :param file: 上传的excel二进制数据
        :param header: # excel表头映射成模型的字段名
        :param serializer: # 序列化器 验证excel中的数据
        :param first_row: # 从哪一行开始获取数据
        :param sheet_no: # 获取excel中哪一个sheet
        :param func: # 处理每行数据的额外函数 比如添加额外数据 或者 数据验证 或者 保存数据
        """
        self.wb = openpyxl.load_workbook(file)
        sheet_names = self.wb.sheetnames
        self.worksheet = self.wb[sheet_names[sheet_no]]
        # 获取总行数
        # self.nrows = self.worksheet.max_row
        columns = self.worksheet.max_column
        excel_header = [self.worksheet.cell(row=first_row, column=i).value for i in range(1, columns + 1)]
        if header:
            self.header = [header.get(h, h) for h in excel_header]
        else:
            self.header = excel_header
        self.first_row = first_row
        self.func = func

    @property
    def save(self) -> List[Any]:
        datas: List[Any] = []
        for row in self.worksheet.iter_rows(min_row=self.first_row + 1):
            row_data = [cell.value for cell in row]
            data = dict(zip(self.header, row_data))
            if self.func:
                data = self.func(data)
            datas.append(data)
        self.wb.close()
        return datas


class DataToExcel:

    def __init__(self, data: DATA = None, headers: Union[dict, list] = None, name: str = 'Sheet1',
                 is_header: bool = False,
                 ext: str = 'xlsx', title: str = None, merge: list = None, beauty: bool = True,
                 cell_width=15, need_header=True, skip_rows=0, func: Callable[[dict], Any] = None):
        """
        :param headers: excel表格头 {'定义的名称'：’字段名',....}
        :param data: 表格数据 [{'字段名':'字段的值’}]
        :param name: 表格的sheet名
        :param is_header: 是否值导出标题头
        :param func: 映射函数 额外处理data数据
        :param ext: 表格格式后缀 xlsx or xls
        :param title: 表格标题名称
        :param merge: 合并单元格cell  例如['A1:A3'] 合并
        :param beauty: 是否美化格式 当数据量大的时候很费时间
        """
        if data:
            first_data = data[0]
            if isinstance(first_data, dict):
                if not headers:
                    self.headers = {k:k for k in first_data.keys()}
                else:
                    if not isinstance(headers, dict):
                        raise ValueError('data是List[dict], header 必须是 字典')
                    else:
                        self.headers = headers
            else:
                if not headers:
                    raise ValueError('data是List[list], header 不能为空')
                else:
                    if not isinstance(headers, list):
                        raise ValueError('data是List[list], header 必须 为 列表')
                    if len(headers) != len(first_data):
                        raise ValueError('data是List[list], header 长度不匹配')
                    self.headers = headers
        else:
            if not headers:
                raise KeyError('header 和 data 不能同时为空')
            else:
                self.headers = headers
        self.data = data
        if isinstance(self.headers, dict):
            self.excel_field_names = list(self.headers.keys())  # 模型所有字段名
        else:
            self.excel_field_names = self.headers
        self.name = name
        self.is_header = is_header
        self.ext = ext
        self.title = title
        self.merge = merge
        self.beauty = beauty
        self.cell_width = cell_width
        self.need_header = need_header
        self.skip_rows = skip_rows
        self.func = func

    @property
    def export(self):

        wb = openpyxl.Workbook()
        ws = wb.active

        if self.title:
            leng = len(self.excel_field_names)
            first_cont = [self.title] + [''] * (leng - 1)
            last_col = get_column_letter(leng)

            ws.append(first_cont)
            ws.merge_cells(f'A1:{last_col}1')
            ws.row_dimensions[1].height = 30
            bold_24_font = Font(name='等线', size=20, italic=False, color='0d0992', bold=True)
            ws['A1'].font = bold_24_font

            headline = 2
        else:
            headline = 1

        if self.need_header:
            ws.append(self.excel_field_names)

        orange_fill = PatternFill(fill_type='solid', fgColor="A6D5B0")
        for ci in range(len(self.excel_field_names)):
            if self.need_header:
                ws.cell(row=headline, column=ci + 1).fill = orange_fill
            col_letter = get_column_letter(ci + 1)
            ws.column_dimensions[col_letter].width = self.cell_width

        if not self.is_header:
            for ind, obj in enumerate(self.data):
                if isinstance(obj, list):
                    data = obj
                elif isinstance(obj, dict):
                    if self.func:
                        obj = self.func(obj)
                    data = [obj.get(v, v) for v in self.headers.values()]
                else:
                    raise ValueError('data 必须是 List[list or dict]')
                if ind >= self.skip_rows:
                    ws.append(data)

            if self.merge:
                for mer in self.merge:
                    ws.merge_cells(mer)

        if self.beauty:
            # 数据量大了后 每一cell去遍历 很花时间  注意取舍
            row_max = ws.max_row
            con_max = ws.max_column
            for xi in range(1, row_max + 1):
                for ji in range(1, con_max + 1):
                    ws.cell(row=xi, column=ji).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=xi, column=ji).border = Border(top=Side(border_style='thin', color=colors.BLACK),
                                                               bottom=Side(border_style='thin', color=colors.BLACK),
                                                               left=Side(border_style='thin', color=colors.BLACK),
                                                               right=Side(border_style='thin', color=colors.BLACK))

        response = file_response(save_virtual_workbook(wb), self.name, self.ext)
        wb.close()
        return response


class Echo:
    """An object that implements just the write method of the file-like
    interface.
    """

    def write(self, value):
        """Write the value by returning it, instead of storing in a buffer."""
        return value


class DataToCSV:

    def __init__(self, data: List[Union[list, dict]], filename: str = 'somefile'):
        self.data = data
        self.filename = filename

    def gen_data(self, writer):
        for index, data in enumerate(self.data):
            if isinstance(data, list):
                yield writer.writerow(data)
            elif isinstance(data, dict):
                if index == 0:
                    yield writer.writerow(list(data.keys()))
                yield writer.writerow(list(data.values()))

    @property
    def export(self):
        pseudo_buffer = Echo()
        writer = csv.writer(pseudo_buffer)
        data = self.gen_data(writer)
        response = StreamingHttpResponse(data, content_type="text/csv")
        response.charset = 'utf-8-sig'
        response['Content-Disposition'] = "attachment;filename={}.csv".format(
            self.filename.encode('utf-8').decode('ISO-8859-1'))
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response

    
import re, stat, mimetypes, os
from pathlib import Path
from django.utils.http import http_date
from django.views.static import was_modified_since
from django.http import FileResponse


# 支持大文件的断点续传（暂停/继续下载）
class BigFileDownload:

    def __init__(self, folder: str, filename: str, request: Optional["HttpRequest"] = None,
                 file_name: Optional[str] = None):
        """
        :param folder: 下载文件所在的文件夹
        :param filename: 下载文件名 必须带上文件后缀名 例如a.zip  a.jpg
        :param request: request对象
        :param file_name: 保存文件名 默认等于 下载文件名
        """
        self.request = request
        self.path = Path(folder) / filename
        if not self.path.exists():
            raise ValueError('文件不存在')
        if file_name is None:
            self.file_name = filename
        self.stat = self.path.stat()

    @property
    def export(self):

        # 判断下载过程中文件是否被修改过
        if not was_modified_since(self.request.META.get('HTTP_IF_MODIFIED_SINCE'),
                                  self.stat.st_mtime, self.stat.st_size):
            raise ValueError('文件已被修改')

        # 获取文件的content_type
        content_type, encoding = mimetypes.guess_type(self.path)
        content_type = content_type or 'application/octet-stream'

        # 计算读取文件的起始位置
        start_bytes = re.search(r'bytes=(\d+)-', self.request.META.get('HTTP_RANGE', ''), re.S)
        start_bytes = int(start_bytes.group(1)) if start_bytes else 0

        # 打开文件并移动下标到起始位置，客户端点击继续下载时，从上次断开的点继续读取
        the_file = self.path.open('rb')
        the_file.seek(start_bytes, os.SEEK_SET)

        # status=200表示下载开始，status=206表示下载暂停后继续，为了兼容火狐浏览器而区分两种状态
        # 关于django的response对象，参考：https://www.cnblogs.com/scolia/p/5635546.html
        # 关于response的状态码，参考：https://www.cnblogs.com/DeasonGuan/articles/Hanami.html
        # FileResponse默认block_size = 4096，因此迭代器每次读取4KB数据
        response = FileResponse(the_file, content_type=content_type, status=206 if start_bytes > 0 else 200,
                                filename=self.file_name)

        # 'Last-Modified'表示文件修改时间，与'HTTP_IF_MODIFIED_SINCE'对应使用，参考：https://www.jianshu.com/p/b4ecca41bbff
        response['Last-Modified'] = http_date(self.stat.st_mtime)

        # 这里'Content-Length'表示剩余待传输的文件字节长度
        if stat.S_ISREG(self.stat.st_mode):
            response['Content-Length'] = self.stat.st_size - start_bytes
        if encoding:
            response['Content-Encoding'] = encoding

        # 'Content-Range'的'/'之前描述响应覆盖的文件字节范围，起始下标为0，'/'之后描述整个文件长度，与'HTTP_RANGE'对应使用
        # 参考：http://liqwei.com/network/protocol/2011/886.shtml
        response['Content-Range'] = 'bytes %s-%s/%s' % (start_bytes, self.stat.st_size - 1, self.stat.st_size)

        # 'Cache-Control'控制浏览器缓存行为，此处禁止浏览器缓存，参考：https://blog.csdn.net/cominglately/article/details/77685214
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response
